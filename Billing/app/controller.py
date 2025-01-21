from flask import jsonify
from sqlalchemy import create_engine
from app import db  # Import the database instance
from openpyxl import load_workbook
from app import create_app  # Import the factory function
from datetime import datetime
import http.client
import json

# Define a Provider model
class Provider(db.Model):
    __tablename__ = 'Provider'  # Explicitly define the table name
    id = db.Column(db.Integer, primary_key=True, autoincrement=True, nullable=False)  # Primary key can't be NULL
    name = db.Column(db.String(255), unique=True, nullable=False)  # Ensure 'name' is not NULL

class Rate(db.Model):
    __tablename__ = 'Rates'  # Explicitly define the table name
    product_id = db.Column(db.String(50), primary_key=True, nullable=False)  # Primary key can't be NULL
    rate = db.Column(db.Integer, default=0, nullable=False)  # Ensure 'rate' is not NULL
    scope = db.Column(db.Integer, db.ForeignKey('Provider.id'), nullable=False)  # Relates to Provider.id, cannot be NULL

class Truck(db.Model):
    __tablename__ = 'Trucks'  # Explicitly define the table name
    id = db.Column(db.String(10), primary_key=True, unique=True, nullable=False)  # Primary key can't be NULL
    provider_id = db.Column(db.Integer, db.ForeignKey('Provider.id'), nullable=False)  # Relates to Provider.id, cannot be NULL

def update_provider_controller(provider_id, new_name):
    """
    Controller function to update an existing provider's name.
    """
    try:
        # Check if the provider exists
        provider = Provider.query.get(provider_id)
        if not provider:
            return f"Provider with ID {provider_id} not found", 404
        
        # Check if the new name already exists for a different provider
        existing_provider = Provider.query.filter(Provider.name == new_name, Provider.id != provider_id).first()
        if existing_provider:
            return f"Provider name '{new_name}' already exists", 400

        # Update the provider's name
        provider.name = new_name
        db.session.commit()

        return f"Provider {provider_id} updated to '{new_name}'", 200
    except Exception as e:
        # Roll back in case of any errors
        db.session.rollback()
        return f"Failed to update provider: {str(e)}", 500

def add_provider(provider_name):

    if Provider.query.filter_by(name=provider_name).first():
        return jsonify({"error": "Provider already exists"}), 409

    new_provider = Provider(name=provider_name)
    db.session.add(new_provider)
    db.session.commit()

    return jsonify({"id": new_provider.id, "name": new_provider.name}), 201

def health_check_controller():
    try:
        # Try querying a known table
        db.session.query(Provider).first()
        return {"status": "OK"}, 200
    except Exception as e:
        print(f"Error: {str(e)}")
        return {"status": "Failure", "message": "Database unavailable"}, 500

def add_truck(license_id, provider_id):
    try:
        # Check if the provider exists
        provider = db.session.query(Provider).get(provider_id)
        if not provider:
            return jsonify({"error": f"Provider with ID {provider_id} not found"}), 404
        
        # Check if the truck already exists
        existing_truck = db.session.query(Truck).filter_by(id=license_id).first()
        if existing_truck:
            return jsonify({"error": f"Truck with license ID {license_id} already exists"}), 400

        # Add the new truck
        new_truck = Truck(id=license_id, provider_id=provider_id)
        db.session.add(new_truck)
        db.session.commit()

        return jsonify({"id": new_truck.id, "provider_id": new_truck.provider_id}), 201
    
    except Exception as e:
        db.session.rollback()  # Rollback in case of any errors during the DB operation.  undo any changes made to the database during the current session
        return jsonify({"error": f"An error occurred while processing the request: {str(e)}"}), 500
    
def update_truck_provider(truck_id, new_provider_id):
    try:
        # Check if the truck exists
        truck = Truck.query.get(truck_id)
        if not truck:
            return jsonify({"error": f"Truck with ID {truck_id} not found"}), 404
        
        # Check if the new provider exists
        provider = Provider.query.get(new_provider_id)
        if not provider:
            return jsonify({"error": f"Provider with ID {new_provider_id} not found"}), 404
        
        # Update the provider_id of the truck
        truck.provider_id = new_provider_id
        db.session.commit()

        return jsonify({"message": f"Truck with ID {truck_id} updated to provider ID {new_provider_id}"}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"An error occurred: {str(e)}"}), 500

def upload_rates_from_excel(file_path):
    try:
        # Load the workbook and the active sheet using openpyxl
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Validate the structure of the sheet (check for required columns)
        headers = [cell.value for cell in sheet[1]]  # Read headers (first row)
        required_columns = {'Product', 'Rate', 'Scope'}
        
        if not required_columns.issubset(headers):
            return jsonify({"error": "Excel file must contain Product, Rate, and Scope columns"}), 400

        # Clear existing rates
        Rate.query.delete()

        # Process the rows (skip the header row, so start from row 2)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            product = str(row[0]).strip()  # Ensure it's a string
            rate = row[1]  # This will be handled directly as a number (no strip needed)
            scope = str(row[2]).strip() if row[2] is not None else None  # Ensure it's a string, handle None
            
            if scope == 'ALL':
                scope = None  # Handle 'ALL' as None
            
            # Create and add the new rate to the database
            new_rate = Rate(product_id=product, rate=rate, scope=scope)
            db.session.add(new_rate)

        # Commit the changes to the database
        db.session.commit()

        return jsonify({"message": "Rates uploaded successfully"}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

    except Exception:
        return "error_fetching_data"  # Generic error if connection fails

SERVICE_HOST = "host.docker.internal"
WEIGHT_SERVICE_PORT = 5000
BILLING_SERVICE_PORT = 5001
BILLING_SERVICE_HOST = "billing-service"

def get_truck_details(truck_id, from_time_str, to_time_str):    
    endpoint = f"/item/{truck_id}?from={from_time_str}&to={to_time_str}"
    return fetch_data(endpoint, SERVICE_HOST, WEIGHT_SERVICE_PORT)

def fetch_data(endpoint, service, port):
    try:
        conn = http.client.HTTPConnection(service, port)
        conn.request("GET", endpoint)
        response = conn.getresponse()

        if response.status == 404:
            return None  # Truck not found
        elif response.status != 200:
            return "error_fetching_data"  # Unexpected error from weight service
        data = json.loads(response.read().decode("utf-8"))
        conn.close()

        return data # Return truck data as received from weight service
        
    except Exception:
        return "error_fetching_data"  # Generic error if connection fails

# Function to get provider details
def get_provider_details(provider_id):
    provider = db.session.query(Provider).get(provider_id)
    if provider:
        print ("prvider name is: ", provider.name)
        return {"id": provider.id, "name": provider.name}
    return None  # Return None if the provider is not found

# Function to get session details for a list of sessions
def get_session_details(session_ids):
    sessions = []
    for session_id in session_ids:
        session_data = fetch_data(f"/session/{session_id}", SERVICE_HOST, WEIGHT_SERVICE_PORT)
        if session_data and "neto" in session_data and session_data["neto"] != "na":
            sessions.append({
                "session_id": session_id, 
                "amount": session_data.get("neto", 0)
            })
    return sessions

# Function to get rates for products
def get_product_rates():
    endpoint = "/rates"
    rates_data = fetch_data(endpoint, BILLING_SERVICE_HOST, BILLING_SERVICE_PORT)

    if rates_data is None or rates_data == "error_fetching_data":
        return {}  
    
    try:
        return {rate["product"]: rate["rate"] for rate in rates_data}
    except Exception as err:
        return {"error": f"Error processing rates data: {str(err)}"}

def get_produce_mapping(provider_id, from_time_str, to_time_str):
    endpoint = f"/weight?t1={from_time_str}&t2={to_time_str}&id={provider_id}"
    weight_data = fetch_data(endpoint , SERVICE_HOST, WEIGHT_SERVICE_PORT)

    try:
        if isinstance(weight_data, str):
            weight_data = json.loads(weight_data)  # Parse JSON if it's a string

        if not isinstance(weight_data, list):
            raise ValueError("Invalid weight data format")

    except (json.JSONDecodeError, ValueError) as e:
        print(f"Error parsing weight data: {e}")
        return {}
    
    produce_mapping = {}
    for entry in weight_data:
        produce_mapping[entry.get("id")] = entry.get("produce", "unknown")

    return produce_mapping

# Main function to get bill details
def get_bill_details(provider_id, from_time_str, to_time_str):
    provider_data = get_provider_details(provider_id)
    if not provider_data:
        return None  # Provider not found

    trucks = Truck.query.filter_by(provider_id=provider_id).all()
    truck_ids = [truck.id for truck in trucks]

    all_sessions = []
    for truck_id in truck_ids:
        truck_data = get_truck_details(truck_id, from_time_str, to_time_str)
        if truck_data and "sessions" in truck_data:
            all_sessions.extend(truck_data["sessions"])
    sessions = get_session_details(all_sessions)

    produce_mapping = get_produce_mapping(provider_id, from_time_str, to_time_str)

    rates = get_product_rates()
    
    product_summary = {}
    for session in sessions:
        session_id = session["session_id"]
        product = produce_mapping.get(session_id, "unknown")
        amount = session["amount"]
        rate = rates.get(product, 0)

        if product not in product_summary:
            product_summary[product] = {"count": 0, "amount": 0, "rate": rate, "pay": 0}

        product_summary[product]["count"] += 1
        product_summary[product]["amount"] += amount
        product_summary[product]["pay"] += amount * rate

    products = [{"product": k, **v} for k, v in product_summary.items()]

    return {
        "id": provider_id,
        "name": provider_data.get("name", f"Provider {provider_id}"),
        "from": from_time_str,
        "to": to_time_str,
        "truckCount": len(truck_ids),
        "sessionCount": len(all_sessions),
        "products": products,
        "total": sum(p["pay"] for p in products)
    }